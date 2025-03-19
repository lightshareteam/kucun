import os
import openpyxl
import pandas as pd
from datetime import datetime, timedelta
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.db.models import Sum, Q, F
from django.utils import timezone
from django.http import HttpResponse, JsonResponse
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill
from io import BytesIO
from .models import Product, Warehouse, StockMovement, IncomingStock, ProductionOrder
from .forms import ProductForm, WarehouseForm, StockMovementForm, IncomingStockForm, ProductionOrderForm

@login_required
def index(request):
    return redirect('inventory:dashboard')

@login_required
def product_list(request):
    # 获取搜索参数
    search_query = request.GET.get('search', '')
    
    # 基础查询集
    products = Product.objects.all()
    
    # 如果有搜索查询，过滤产品
    if search_query:
        products = products.filter(
            Q(sku__icontains=search_query) |  # 店铺SKU
            Q(name__icontains=search_query)    # 产品SKU
        )
    
    warehouses = Warehouse.objects.all()
    
    # 获取亚特兰大和加州仓库
    ga_warehouse = None
    ca_warehouse = None
    for warehouse in warehouses:
        if warehouse.code == 'GA':
            ga_warehouse = warehouse
        elif warehouse.code == 'CA':
            ca_warehouse = warehouse
    
    # 预先计算每个产品在每个仓库的库存
    product_warehouse_stock = {}
    # 预先计算每个产品在每个仓库的在途数量
    product_incoming_stock = {}
    # 预先计算每个产品的生产订单剩余数量
    product_production_orders = {}
    
    # 获取所有待入库的入库途中产品
    incoming_stocks = IncomingStock.objects.filter(status='待入库')
    
    # 获取所有未完成的生产订单
    production_orders = ProductionOrder.objects.filter(remaining_quantity__gt=0)
    
    # 初始化产品在途数据字典和生产订单数据字典
    for product in products:
        product_warehouse_stock[product.id] = {}
        product_incoming_stock[product.id] = {
            'GA': 0,
            'CA': 0,
            'total': 0
        }
        product_production_orders[product.id] = 0
        
        # 计算每个仓库的库存
        for warehouse in warehouses:
            product_warehouse_stock[product.id][warehouse.id] = product.get_stock_by_warehouse(warehouse)
    
    # 计算每个产品在GA和CA的在途数量
    for incoming in incoming_stocks:
        if incoming.product.id in product_incoming_stock:
            if incoming.warehouse.code == 'GA':
                product_incoming_stock[incoming.product.id]['GA'] += incoming.quantity
            elif incoming.warehouse.code == 'CA':
                product_incoming_stock[incoming.product.id]['CA'] += incoming.quantity
            product_incoming_stock[incoming.product.id]['total'] += incoming.quantity
    
    # 计算每个产品的生产订单剩余数量
    for order in production_orders:
        if order.product.id in product_production_orders:
            product_production_orders[order.product.id] += order.remaining_quantity
    
    return render(request, 'inventory/product_list.html', {
        'products': products,
        'warehouses': warehouses,
        'product_warehouse_stock': product_warehouse_stock,
        'product_incoming_stock': product_incoming_stock,
        'product_production_orders': product_production_orders,
        'search_query': search_query,  # 传递搜索查询到模板
        'ga_warehouse': ga_warehouse,
        'ca_warehouse': ca_warehouse,
    })

@login_required
def product_create(request):
    if request.method == 'POST':
        form = ProductForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, '产品创建成功！')
            return redirect('inventory:product_list')
    else:
        form = ProductForm()
    return render(request, 'inventory/product_form.html', {'form': form})

@login_required
def product_edit(request, pk):
    product = get_object_or_404(Product, pk=pk)
    if request.method == 'POST':
        form = ProductForm(request.POST, instance=product)
        if form.is_valid():
            form.save()
            messages.success(request, '产品更新成功！')
            return redirect('inventory:product_list')
    else:
        form = ProductForm(instance=product)
    return render(request, 'inventory/product_form.html', {'form': form})

@login_required
def product_delete(request, pk):
    product = get_object_or_404(Product, pk=pk)
    product_name = product.name
    
    # 检查是否有关联的库存变动记录
    stock_movements = StockMovement.objects.filter(product=product)
    if stock_movements.exists():
        messages.error(request, f'无法删除产品 "{product_name}"，因为它有关联的库存变动记录。')
    else:
        product.delete()
        messages.success(request, f'产品 "{product_name}" 已成功删除。')
    
    return redirect('inventory:product_list')

@login_required
def product_batch_delete(request):
    if request.method == 'POST':
        selected_ids = request.POST.getlist('selected_products')
        if not selected_ids:
            messages.error(request, '未选择任何产品。')
            return redirect('inventory:product_list')
        
        success_count = 0
        error_count = 0
        
        for product_id in selected_ids:
            try:
                product = Product.objects.get(pk=product_id)
                
                # 检查是否有关联的库存变动记录
                stock_movements = StockMovement.objects.filter(product=product)
                if stock_movements.exists():
                    error_count += 1
                else:
                    product.delete()
                    success_count += 1
            except Product.DoesNotExist:
                error_count += 1
        
        if success_count > 0:
            messages.success(request, f'成功删除 {success_count} 个产品。')
        if error_count > 0:
            messages.error(request, f'{error_count} 个产品无法删除，因为它们有关联的库存变动记录。')
    
    return redirect('inventory:product_list')

@login_required
def warehouse_list(request):
    warehouses = Warehouse.objects.all()
    return render(request, 'inventory/warehouse_list.html', {'warehouses': warehouses})

@login_required
def warehouse_create(request):
    if request.method == 'POST':
        form = WarehouseForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, '仓库创建成功！')
            return redirect('inventory:warehouse_list')
    else:
        form = WarehouseForm()
    return render(request, 'inventory/warehouse_form.html', {'form': form})

@login_required
def warehouse_edit(request, pk):
    warehouse = get_object_or_404(Warehouse, pk=pk)
    if request.method == 'POST':
        form = WarehouseForm(request.POST, instance=warehouse)
        if form.is_valid():
            form.save()
            messages.success(request, '仓库更新成功！')
            return redirect('inventory:warehouse_list')
    else:
        form = WarehouseForm(instance=warehouse)
    return render(request, 'inventory/warehouse_form.html', {'form': form})

@login_required
def warehouse_delete(request, pk):
    warehouse = get_object_or_404(Warehouse, pk=pk)
    warehouse_name = warehouse.name
    
    # 检查是否有关联的库存变动记录
    stock_movements = StockMovement.objects.filter(warehouse=warehouse)
    if stock_movements.exists():
        messages.error(request, f'无法删除仓库 "{warehouse_name}"，因为它有关联的库存变动记录。')
    else:
        warehouse.delete()
        messages.success(request, f'仓库 "{warehouse_name}" 已成功删除。')
    
    return redirect('inventory:warehouse_list')

@login_required
def warehouse_batch_delete(request):
    if request.method == 'POST':
        selected_ids = request.POST.getlist('selected_warehouses')
        if not selected_ids:
            messages.error(request, '未选择任何仓库。')
            return redirect('inventory:warehouse_list')
        
        success_count = 0
        error_count = 0
        
        for warehouse_id in selected_ids:
            try:
                warehouse = Warehouse.objects.get(pk=warehouse_id)
                
                # 检查是否有关联的库存变动记录
                stock_movements = StockMovement.objects.filter(warehouse=warehouse)
                if stock_movements.exists():
                    error_count += 1
                else:
                    warehouse.delete()
                    success_count += 1
            except Warehouse.DoesNotExist:
                error_count += 1
        
        if success_count > 0:
            messages.success(request, f'成功删除 {success_count} 个仓库。')
        if error_count > 0:
            messages.error(request, f'{error_count} 个仓库无法删除，因为它们有关联的库存变动记录。')
    
    return redirect('inventory:warehouse_list')

@login_required
def stock_movement_list(request):
    movements = StockMovement.objects.all().order_by('-date', '-created_at')
    return render(request, 'inventory/stock_movement_list.html', {'movements': movements})

@login_required
def stock_movement_create(request):
    # 获取URL参数中的产品ID
    product_id = request.GET.get('product')
    initial_data = {'date': timezone.now().date()}
    
    # 如果有产品ID参数，则预设产品和入库类型
    if product_id:
        try:
            product = Product.objects.get(id=product_id)
            initial_data.update({
                'product': product,
                'movement_type': 'IN'  # 默认为入库
            })
        except Product.DoesNotExist:
            pass
    
    if request.method == 'POST':
        form = StockMovementForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, '库存变动记录创建成功！')
            return redirect('inventory:stock_movement_list')
    else:
        form = StockMovementForm(initial=initial_data)
    
    return render(request, 'inventory/stock_movement_form.html', {'form': form})

@login_required
def stock_movement_delete(request, pk):
    movement = get_object_or_404(StockMovement, pk=pk)
    product_name = movement.product.name
    movement_type = '入库' if movement.movement_type == 'IN' else '出库'
    quantity = movement.quantity
    
    movement.delete()
    messages.success(request, f'已删除 {product_name} 的{movement_type}记录（数量：{quantity}）。')
    
    return redirect('inventory:stock_movement_list')

@login_required
def stock_movement_batch_delete(request):
    if request.method == 'POST':
        selected_ids = request.POST.getlist('selected_movements')
        if not selected_ids:
            messages.error(request, '未选择任何库存变动记录。')
            return redirect('inventory:stock_movement_list')
        
        deleted_count = 0
        
        for movement_id in selected_ids:
            try:
                movement = StockMovement.objects.get(pk=movement_id)
                movement.delete()
                deleted_count += 1
            except StockMovement.DoesNotExist:
                pass
        
        if deleted_count > 0:
            messages.success(request, f'成功删除 {deleted_count} 条库存变动记录。')
    
    return redirect('inventory:stock_movement_list')

@login_required
def dashboard(request):
    """仪表盘视图"""
    # 获取产品总数
    total_products = Product.objects.count()
    
    # 获取仓库总数
    total_warehouses = Warehouse.objects.count()
    
    # 获取入库途中产品总数
    total_incoming_stock = IncomingStock.objects.filter(status='待入库').count()
    
    # 获取所有产品的产品SKU列表（name字段）
    all_product_skus = Product.objects.values_list('name', flat=True).distinct()
    
    # 获取低库存产品
    low_stock_products = []
    
    for product_sku in all_product_skus:
        # 获取相同产品SKU的所有产品
        sku_products = Product.objects.filter(name=product_sku)
        
        # 计算这些产品的总库存
        total_stock = 0
        for product in sku_products:
            # 使用产品的total_stock属性计算库存
            total_stock += product.total_stock
        
        # 获取第一个产品的库存阈值作为参考
        if sku_products.exists():
            reference_product = sku_products.first()
            threshold = reference_product.low_stock_threshold
            
            # 如果总库存低于阈值，添加到低库存列表
            if total_stock < threshold and threshold > 0:
                low_stock_products.append({
                    'id': reference_product.id,
                    'sku': product_sku,  # 使用产品SKU（name字段）
                    'current_stock': total_stock,
                    'low_stock_threshold': threshold
                })
    
    # 按库存量排序
    low_stock_products = sorted(low_stock_products, key=lambda x: x['current_stock'])
    
    # 获取各仓库的库存统计
    warehouses = Warehouse.objects.all()
    warehouse_stats = []
    
    for warehouse in warehouses:
        # 计算该仓库的总库存数量
        total_stock = StockMovement.objects.filter(warehouse=warehouse).aggregate(
            total=Sum('quantity', filter=Q(movement_type='IN')) - 
                  Sum('quantity', filter=Q(movement_type='OUT'))
        )['total'] or 0
        
        # 计算该仓库的入库途中产品数量
        incoming_count = IncomingStock.objects.filter(
            warehouse=warehouse,
            status='待入库'
        ).aggregate(total=Sum('quantity'))['total'] or 0
        
        warehouse_stats.append({
            'warehouse': warehouse,
            'total_stock': total_stock,
            'incoming_count': incoming_count
        })
    
    context = {
        'total_products': total_products,
        'total_warehouses': total_warehouses,
        'total_incoming_stock': total_incoming_stock,
        'low_stock_products': low_stock_products,
        'warehouse_stats': warehouse_stats,
    }
    
    return render(request, 'inventory/dashboard.html', context)

@login_required
def download_stock_movement_template(request):
    # 创建一个新的工作簿
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "库存变动导入模板"
    
    # 设置表头
    headers = ['日期(YYYY-MM-DD)', '仓库代码', '店铺SKU', '数量', '类型(IN/OUT)', '备注(可选)']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col)
        cell.value = header
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
    
    # 添加示例数据
    example_data = [
        ['2025-03-07', 'GA', 'SHOP001', '100', 'IN', '亚特兰大仓库入库'],
        ['2025-03-07', 'CA', 'SHOP002', '50', 'OUT', '加州仓库发货']
    ]
    for row, data in enumerate(example_data, 2):
        for col, value in enumerate(data, 1):
            ws.cell(row=row, column=col, value=value)
    
    # 设置列宽
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 15
    
    # 保存到内存中
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    # 返回文件
    response = HttpResponse(buffer.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=stock_movement_template.xlsx'
    return response

@login_required
def import_stock_movements(request):
    if request.method == 'POST' and request.FILES.get('file'):
        try:
            excel_file = request.FILES['file']
            df = pd.read_excel(excel_file)
            
            # 验证必要的列是否存在
            required_columns = ['日期(YYYY-MM-DD)', '仓库代码', '店铺SKU', '数量', '类型(IN/OUT)']
            for col in required_columns:
                if col not in df.columns:
                    messages.error(request, f'缺少必要的列: {col}')
                    return redirect('inventory:stock_movement_list')
            
            success_count = 0
            error_count = 0
            error_messages = []
            
            for index, row in df.iterrows():
                try:
                    # 获取仓库
                    try:
                        warehouse = Warehouse.objects.get(code=row['仓库代码'])
                    except Warehouse.DoesNotExist:
                        error_messages.append(f'第{index+2}行: 仓库代码 {row["仓库代码"]} 不存在')
                        error_count += 1
                        continue
                    
                    # 获取产品
                    try:
                        product = Product.objects.get(sku=row['店铺SKU'])
                    except Product.DoesNotExist:
                        error_messages.append(f'第{index+2}行: 店铺SKU {row["店铺SKU"]} 不存在')
                        error_count += 1
                        continue
                    
                    # 验证并转换日期
                    try:
                        date = pd.to_datetime(row['日期(YYYY-MM-DD)']).date()
                    except:
                        error_messages.append(f'第{index+2}行: 日期格式错误')
                        error_count += 1
                        continue
                    
                    # 验证数量
                    try:
                        quantity = int(row['数量'])
                        if quantity <= 0:
                            raise ValueError
                    except:
                        error_messages.append(f'第{index+2}行: 数量必须为正整数')
                        error_count += 1
                        continue
                    
                    # 验证类型
                    movement_type = str(row['类型(IN/OUT)']).strip().upper()
                    if movement_type not in ['IN', 'OUT']:
                        error_messages.append(f'第{index+2}行: 类型必须为 IN 或 OUT')
                        error_count += 1
                        continue
                    
                    # 获取备注
                    notes = str(row.get('备注(可选)', '')) if pd.notna(row.get('备注(可选)', '')) else ''
                    
                    # 创建库存变动记录
                    StockMovement.objects.create(
                        warehouse=warehouse,
                        product=product,
                        date=date,
                        quantity=quantity,
                        movement_type=movement_type,
                        notes=notes
                    )
                    success_count += 1
                    
                except Exception as e:
                    error_messages.append(f'第{index+2}行: 处理出错 - {str(e)}')
                    error_count += 1
            
            # 显示导入结果
            if success_count > 0:
                messages.success(request, f'成功导入 {success_count} 条记录')
            if error_count > 0:
                messages.error(request, f'导入失败 {error_count} 条记录')
                for msg in error_messages[:5]:  # 只显示前5条错误信息
                    messages.warning(request, msg)
                if len(error_messages) > 5:
                    messages.warning(request, '...')
            
        except Exception as e:
            messages.error(request, f'文件处理出错: {str(e)}')
        
        return redirect('inventory:stock_movement_list')
    
    return render(request, 'inventory/stock_movement_import.html')

@login_required
def download_product_template(request):
    # 创建一个新的工作簿
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "产品导入模板"
    
    # 设置列标题
    headers = ['店铺SKU', 'FNSKU', '产品SKU', '重量(lb)', '长度(inch)', '宽度(inch)', '高度(inch)', '亚特兰大仓库存', '加州仓库存', '总库存(参考)']
    
    # 写入标题行
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = Font(bold=True)
    
    # 添加示例数据
    example_data = [
        'EXAMPLE-SKU', 'X00EXAMPLE', 'EXAMPLE-SKU', 5, 10, 10, 10, 100, 50, '=SUM(H2:I2)'
    ]
    
    for col_num, value in enumerate(example_data, 1):
        ws.cell(row=2, column=col_num, value=value)
    
    # 设置列宽
    for col_num in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col_num)].width = 15
    
    # 保存到内存中
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    # 创建响应
    response = HttpResponse(
        buffer.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=product_import_template.xlsx'
    
    return response

@login_required
def import_products(request):
    if request.method == 'POST' and request.FILES.get('file'):
        try:
            excel_file = request.FILES['file']
            df = pd.read_excel(excel_file)
            
            # 打印列名，用于调试
            print("Excel文件列名:", df.columns.tolist())
            
            # 验证必要的列是否存在
            required_columns = ['店铺SKU', '重量(lb)', '长度(inch)', '宽度(inch)', '高度(inch)']
            for col in required_columns:
                if col not in df.columns:
                    messages.error(request, f'缺少必要的列: {col}')
                    return redirect('inventory:product_list')
            
            # 获取仓库
            atlanta_warehouse = Warehouse.objects.get(code='GA')
            california_warehouse = Warehouse.objects.get(code='CA')
            
            success_count = 0
            error_count = 0
            error_messages = []
            stock_added_count = 0
            
            for index, row in df.iterrows():
                try:
                    # 打印当前行数据，用于调试
                    print(f"处理第{index+2}行: {dict(row)}")
                    
                    # 验证店铺SKU
                    sku = str(row['店铺SKU']).strip()
                    if not sku:
                        error_messages.append(f'第{index+2}行: 店铺SKU不能为空')
                        error_count += 1
                        continue
                    
                    # 检查店铺SKU是否已存在
                    existing_product = Product.objects.filter(sku=sku).first()
                    
                    # 获取产品SKU（如果存在）
                    product_sku = str(row.get('产品SKU', '')) if pd.notna(row.get('产品SKU', '')) else sku
                    
                    # 验证数值字段
                    try:
                        weight = float(row['重量(lb)'])
                        length = float(row['长度(inch)'])
                        width = float(row['宽度(inch)'])
                        height = float(row['高度(inch)'])
                        
                        if weight <= 0 or length <= 0 or width <= 0 or height <= 0:
                            error_messages.append(f'第{index+2}行: 重量和尺寸必须为正数')
                            error_count += 1
                            continue
                    except:
                        error_messages.append(f'第{index+2}行: 重量和尺寸必须为数字')
                        error_count += 1
                        continue
                    
                    # 获取FNSKU（可选）
                    fnsku = str(row.get('FNSKU', '')) if pd.notna(row.get('FNSKU', '')) else ''
                    
                    # 获取库存数量（可选）
                    atlanta_stock = 0
                    california_stock = 0
                    
                    # 检查亚特兰大仓库存列是否存在
                    if '亚特兰大仓库存' in df.columns:
                        try:
                            if pd.notna(row['亚特兰大仓库存']):
                                atlanta_stock = int(float(row['亚特兰大仓库存']))
                                print(f"亚特兰大仓库存: {atlanta_stock}")
                                if atlanta_stock < 0:
                                    error_messages.append(f'第{index+2}行: 亚特兰大仓库存不能为负数')
                                    atlanta_stock = 0
                        except (ValueError, TypeError) as e:
                            print(f"亚特兰大仓库存转换错误: {e}")
                            error_messages.append(f'第{index+2}行: 亚特兰大仓库存必须为整数')
                            atlanta_stock = 0
                    else:
                        print("Excel文件中没有'亚特兰大仓库存'列")
                    
                    # 检查加州仓库存列是否存在
                    if '加州仓库存' in df.columns:
                        try:
                            if pd.notna(row['加州仓库存']):
                                california_stock = int(float(row['加州仓库存']))
                                print(f"加州仓库存: {california_stock}")
                                if california_stock < 0:
                                    error_messages.append(f'第{index+2}行: 加州仓库存不能为负数')
                                    california_stock = 0
                        except (ValueError, TypeError) as e:
                            print(f"加州仓库存转换错误: {e}")
                            error_messages.append(f'第{index+2}行: 加州仓库存必须为整数')
                            california_stock = 0
                    else:
                        print("Excel文件中没有'加州仓库存'列")
                    
                    # 创建或更新产品
                    if existing_product:
                        existing_product.fnsku = fnsku
                        existing_product.name = product_sku
                        existing_product.weight = weight
                        existing_product.length = length
                        existing_product.width = width
                        existing_product.height = height
                        existing_product.save()
                        product = existing_product
                        print(f"更新产品: {product.sku} (ID: {product.id})")
                    else:
                        product = Product.objects.create(
                            sku=sku,
                            fnsku=fnsku,
                            name=product_sku,  # 使用产品SKU作为产品名称
                            weight=weight,
                            length=length,
                            width=width,
                            height=height
                        )
                        print(f"创建产品: {product.sku} (ID: {product.id})")
                    
                    # 处理库存数据
                    today = timezone.now().date()
                    
                    # 亚特兰大仓库库存
                    if atlanta_stock > 0:
                        stock_movement = StockMovement.objects.create(
                            product=product,
                            warehouse=atlanta_warehouse,
                            movement_type='IN',
                            quantity=atlanta_stock,
                            date=today,
                            notes='导入初始库存'
                        )
                        print(f"创建亚特兰大库存变动记录: ID={stock_movement.id}, 产品={product.sku}, 仓库={atlanta_warehouse.name}, 数量={atlanta_stock}")
                        stock_added_count += 1
                        messages.info(request, f'产品 {product.sku} 已添加 {atlanta_stock} 件库存到亚特兰大仓库')
                    
                    # 加州仓库库存
                    if california_stock > 0:
                        stock_movement = StockMovement.objects.create(
                            product=product,
                            warehouse=california_warehouse,
                            movement_type='IN',
                            quantity=california_stock,
                            date=today,
                            notes='导入初始库存'
                        )
                        print(f"创建加州库存变动记录: ID={stock_movement.id}, 产品={product.sku}, 仓库={california_warehouse.name}, 数量={california_stock}")
                        stock_added_count += 1
                        messages.info(request, f'产品 {product.sku} 已添加 {california_stock} 件库存到加州仓库')
                    
                    success_count += 1
                    
                except Exception as e:
                    print(f"处理第{index+2}行时出错: {str(e)}")
                    error_messages.append(f'第{index+2}行: 处理出错 - {str(e)}')
                    error_count += 1
            
            # 显示导入结果
            if success_count > 0:
                messages.success(request, f'成功导入 {success_count} 条产品记录')
            if stock_added_count > 0:
                messages.success(request, f'成功添加 {stock_added_count} 条库存记录')
            if error_count > 0:
                messages.error(request, f'导入失败 {error_count} 条记录')
                for msg in error_messages[:5]:  # 只显示前5条错误信息
                    messages.warning(request, msg)
                if len(error_messages) > 5:
                    messages.warning(request, '...')
            
        except Exception as e:
            print(f"文件处理出错: {str(e)}")
            messages.error(request, f'文件处理出错: {str(e)}')
        
        return redirect('inventory:product_list')
    
    return render(request, 'inventory/product_import.html')

@login_required
def export_products(request):
    # 创建一个新的工作簿
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "产品列表"
    
    # 获取仓库
    atlanta_warehouse = Warehouse.objects.get(code='GA')
    california_warehouse = Warehouse.objects.get(code='CA')
    
    # 设置表头 - 确保与导入模板一致
    headers = ['店铺SKU', 'FNSKU', '产品SKU', '重量(lb)', '长度(inch)', '宽度(inch)', '高度(inch)', '亚特兰大仓库存', '加州仓库存', '总库存(参考)']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col)
        cell.value = header
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
    
    # 添加产品数据
    products = Product.objects.all()
    for row, product in enumerate(products, 2):
        # 获取各仓库库存
        atlanta_stock = product.get_stock_by_warehouse(atlanta_warehouse)
        california_stock = product.get_stock_by_warehouse(california_warehouse)
        total_stock = product.total_stock
        
        ws.cell(row=row, column=1, value=product.sku)
        ws.cell(row=row, column=2, value=product.fnsku)
        ws.cell(row=row, column=3, value=product.name)
        ws.cell(row=row, column=4, value=float(product.weight))
        ws.cell(row=row, column=5, value=float(product.length))
        ws.cell(row=row, column=6, value=float(product.width))
        ws.cell(row=row, column=7, value=float(product.height))
        ws.cell(row=row, column=8, value=atlanta_stock)
        ws.cell(row=row, column=9, value=california_stock)
        ws.cell(row=row, column=10, value=total_stock)
    
    # 设置列宽
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 15
    
    # 保存到内存中
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    # 返回文件
    response = HttpResponse(buffer.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=products_{}.xlsx'.format(datetime.now().strftime('%Y%m%d%H%M%S'))
    return response

@login_required
def historical_stock(request):
    """历史库存查询视图"""
    products = Product.objects.all()
    warehouses = Warehouse.objects.all()
    
    # 默认日期为今天
    selected_date = timezone.now().date()
    
    # 如果有日期参数，使用参数日期
    if request.GET.get('date'):
        try:
            selected_date = datetime.strptime(request.GET.get('date'), '%Y-%m-%d').date()
        except ValueError:
            messages.error(request, '日期格式错误，请使用YYYY-MM-DD格式')
    
    # 计算历史库存
    product_warehouse_stock = {}
    for product in products:
        product_warehouse_stock[product.id] = {}
        for warehouse in warehouses:
            # 计算截至所选日期的库存
            total_in = StockMovement.objects.filter(
                product=product,
                warehouse=warehouse,
                movement_type='IN',
                date__lte=selected_date
            ).aggregate(total=Sum('quantity'))['total'] or 0

            total_out = StockMovement.objects.filter(
                product=product,
                warehouse=warehouse,
                movement_type='OUT',
                date__lte=selected_date
            ).aggregate(total=Sum('quantity'))['total'] or 0

            product_warehouse_stock[product.id][warehouse.id] = total_in - total_out
    
    # 计算每个产品的总库存
    product_total_stock = {}
    for product in products:
        total_in = StockMovement.objects.filter(
            product=product,
            movement_type='IN',
            date__lte=selected_date
        ).aggregate(total=Sum('quantity'))['total'] or 0

        total_out = StockMovement.objects.filter(
            product=product,
            movement_type='OUT',
            date__lte=selected_date
        ).aggregate(total=Sum('quantity'))['total'] or 0

        product_total_stock[product.id] = total_in - total_out
    
    # 生成日期选择器的日期范围
    # 获取最早的库存记录日期
    earliest_movement = StockMovement.objects.order_by('date').first()
    earliest_date = earliest_movement.date if earliest_movement else timezone.now().date()
    
    # 生成日期列表，从最早日期到今天
    today = timezone.now().date()
    date_range = []
    current_date = earliest_date
    while current_date <= today:
        date_range.append(current_date)
        current_date += timedelta(days=1)
    
    return render(request, 'inventory/historical_stock.html', {
        'products': products,
        'warehouses': warehouses,
        'product_warehouse_stock': product_warehouse_stock,
        'product_total_stock': product_total_stock,
        'selected_date': selected_date,
        'date_range': date_range,
    })

@login_required
def incoming_stock_list(request):
    """入库途中产品列表视图"""
    incoming_stock_list = IncomingStock.objects.all().order_by('-expected_arrival_date')
    return render(request, 'inventory/incoming_stock_list.html', {
        'incoming_stock_list': incoming_stock_list,
    })

@login_required
def incoming_stock_create(request):
    """创建入库途中产品记录并自动扣减生产订单数量"""
    if request.method == 'POST':
        form = IncomingStockForm(request.POST)
        print(f"表单提交数据: {request.POST}")
        
        if form.is_valid():
            print("表单验证成功")
            # 获取表单数据但不立即保存
            incoming_stock = form.save(commit=False)
            
            # 打印调试信息
            print(f"创建入库途中产品记录: 产品={incoming_stock.product.sku}, 数量={incoming_stock.quantity}, 状态={incoming_stock.status}")
            
            # 确保状态为待入库
            if incoming_stock.status != '待入库':
                print(f"手动设置状态为待入库，原状态为{incoming_stock.status}")
                incoming_stock.status = '待入库'
            
            # 保存入库途中产品记录
            incoming_stock.save()
            
            # 检查是否有相应的生产订单
            production_orders = ProductionOrder.objects.filter(
                product=incoming_stock.product,
                remaining_quantity__gt=0
            )
            
            if not production_orders.exists():
                print(f"警告：产品{incoming_stock.product.sku}没有可用的生产订单，无法扣减剩余数量")
                messages.warning(request, f'入库途中产品记录创建成功，但没有找到可用的生产订单，无法扣减剩余数量。')
                return redirect('inventory:incoming_stock_list')
            
            # 自动扣减生产订单数量
            print(f"状态为待入库，准备扣减生产订单数量")
            try:
                # 使用通用的adjust_production_orders函数扣减生产订单数量
                orders_updated = adjust_production_orders(incoming_stock.product, incoming_stock.quantity)
                print(f"扣减生产订单数量完成，更新了{orders_updated}个订单")
                
                if orders_updated > 0:
                    messages.success(request, f'入库途中产品记录创建成功，并自动扣减了 {orders_updated} 个生产订单的数量。')
                else:
                    messages.success(request, '入库途中产品记录创建成功，但没有扣减任何生产订单的数量。')
            except Exception as e:
                print(f"扣减生产订单数量时出错: {str(e)}")
                messages.error(request, f'入库途中产品记录创建成功，但扣减生产订单数量时出错: {str(e)}')
            
            return redirect('inventory:incoming_stock_list')
        else:
            print(f"表单验证失败: {form.errors}")
            messages.error(request, f'表单验证失败，请检查输入: {form.errors}')
    else:
        form = IncomingStockForm(initial={'expected_arrival_date': timezone.now().date(), 'status': '待入库'})
    
    return render(request, 'inventory/incoming_stock_form.html', {'form': form})

@login_required
def incoming_stock_edit(request, pk):
    """编辑入库途中产品记录并自动调整生产订单数量"""
    incoming_stock = get_object_or_404(IncomingStock, pk=pk)
    original_quantity = incoming_stock.quantity
    original_status = incoming_stock.status
    
    if request.method == 'POST':
        form = IncomingStockForm(request.POST, instance=incoming_stock)
        if form.is_valid():
            # 获取新的数量和状态
            new_quantity = form.cleaned_data['quantity']
            new_status = form.cleaned_data['status']
            
            # 保存更新后的入库途中产品记录
            updated_incoming_stock = form.save()
            
            try:
                # 处理生产订单数量调整
                quantity_diff = 0
                
                # 如果原状态是非待入库，新状态是待入库，需要扣减生产订单
                if original_status != '待入库' and new_status == '待入库':
                    quantity_diff = new_quantity
                # 如果原状态是待入库，新状态是非待入库，需要增加生产订单
                elif original_status == '待入库' and new_status != '待入库':
                    quantity_diff = -original_quantity
                # 如果状态都是待入库，但数量有变化，需要调整差额
                elif original_status == '待入库' and new_status == '待入库' and original_quantity != new_quantity:
                    quantity_diff = new_quantity - original_quantity
                
                if quantity_diff != 0:
                    orders_updated = adjust_production_orders(updated_incoming_stock.product, quantity_diff)
                    if orders_updated > 0:
                        if quantity_diff > 0:
                            messages.success(request, f'入库途中产品记录更新成功，并扣减了 {orders_updated} 个生产订单的数量。')
                        else:
                            messages.success(request, f'入库途中产品记录更新成功，并恢复了 {orders_updated} 个生产订单的数量。')
                    else:
                        messages.success(request, '入库途中产品记录更新成功。')
                else:
                    messages.success(request, '入库途中产品记录更新成功。')
            except Exception as e:
                messages.error(request, f'调整生产订单数量时出错: {str(e)}')
            
            return redirect('inventory:incoming_stock_list')
    else:
        form = IncomingStockForm(instance=incoming_stock)
    
    return render(request, 'inventory/incoming_stock_form.html', {'form': form})

def adjust_production_orders(product, quantity_diff):
    """调整生产订单数量
    
    参数:
        product: 产品对象
        quantity_diff: 数量差异，正数表示扣减，负数表示增加
    
    返回:
        更新的订单数量
    """
    print(f"调用adjust_production_orders: 产品={product.sku}, 数量差异={quantity_diff}")
    orders_updated = 0
    
    if quantity_diff > 0:
        # 扣减生产订单数量
        # 获取该产品的所有生产订单，按创建时间排序（先处理早期订单）
        production_orders = ProductionOrder.objects.filter(
            product=product,
            remaining_quantity__gt=0
        ).order_by('created_at')
        
        print(f"找到{production_orders.count()}个有剩余数量的生产订单")
        
        deducted_quantity = quantity_diff
        for order in production_orders:
            if deducted_quantity <= 0:
                break
            
            # 计算可以从当前订单扣减的数量
            deduct_from_order = min(deducted_quantity, order.remaining_quantity)
            
            if deduct_from_order > 0:
                print(f"从订单{order.order_number}扣减{deduct_from_order}，原剩余数量={order.remaining_quantity}")
                # 更新订单剩余数量
                order.remaining_quantity -= deduct_from_order
                order.save()
                print(f"订单{order.order_number}更新后剩余数量={order.remaining_quantity}")
                deducted_quantity -= deduct_from_order
                orders_updated += 1
    
    elif quantity_diff < 0:
        # 增加生产订单数量（恢复）
        # 获取该产品的所有生产订单，按创建时间倒序排序（先处理最近的订单）
        production_orders = ProductionOrder.objects.filter(
            product=product
        ).order_by('-created_at')
        
        add_quantity = -quantity_diff  # 转为正数
        for order in production_orders:
            if add_quantity <= 0:
                break
            
            # 计算可以恢复到当前订单的数量（不超过原始订单数量）
            add_to_order = min(add_quantity, order.quantity - order.remaining_quantity)
            
            if add_to_order > 0:
                # 更新订单剩余数量
                order.remaining_quantity += add_to_order
                order.save()
                add_quantity -= add_to_order
                orders_updated += 1
    
    return orders_updated

@login_required
def incoming_stock_delete(request, pk):
    """删除入库途中产品记录并恢复生产订单数量"""
    incoming_stock = get_object_or_404(IncomingStock, pk=pk)
    
    if request.method == 'POST':
        # 如果是待入库状态，需要恢复生产订单数量
        if incoming_stock.status == '待入库':
            try:
                # 恢复生产订单数量（传入负数表示增加）
                orders_updated = adjust_production_orders(incoming_stock.product, -incoming_stock.quantity)
                
                # 删除入库途中产品记录
                incoming_stock.delete()
                
                if orders_updated > 0:
                    messages.success(request, f'入库途中产品记录已成功删除，并恢复了 {orders_updated} 个生产订单的数量。')
                else:
                    messages.success(request, '入库途中产品记录已成功删除。')
            except Exception as e:
                messages.error(request, f'恢复生产订单数量时出错: {str(e)}')
                return redirect('inventory:incoming_stock_list')
        else:
            # 直接删除
            incoming_stock.delete()
            messages.success(request, '入库途中产品记录已成功删除！')
        
        return redirect('inventory:incoming_stock_list')
    
    return render(request, 'inventory/incoming_stock_confirm_delete.html', {'incoming_stock': incoming_stock})

@login_required
def incoming_stock_batch_delete(request):
    """批量删除入库途中产品记录"""
    if request.method == 'POST':
        selected_ids = request.POST.getlist('selected_items')
        print(f"接收到批量删除请求，选中的ID: {selected_ids}")
        
        if not selected_ids:
            messages.error(request, '未选择任何入库途中产品记录。')
            return redirect('inventory:incoming_stock_list')
        
        deleted_count = 0
        error_count = 0
        
        for item_id in selected_ids:
            try:
                item = IncomingStock.objects.get(pk=item_id)
                item_status = item.status
                item_sku = item.product.sku if item.product else "未知产品"
                item_warehouse = item.warehouse.code if item.warehouse else "未知仓库"
                
                item.delete()
                deleted_count += 1
                print(f"成功删除入库途中产品记录 ID={item_id}, 产品={item_sku}, 仓库={item_warehouse}, 状态={item_status}")
            except IncomingStock.DoesNotExist:
                error_count += 1
                print(f"入库途中产品记录不存在 ID={item_id}")
            except Exception as e:
                error_count += 1
                print(f"删除入库途中产品记录时出错 ID={item_id}, 错误: {str(e)}")
        
        if deleted_count > 0:
            messages.success(request, f'成功删除 {deleted_count} 条入库途中产品记录。')
        else:
            messages.warning(request, '没有删除任何入库途中产品记录。')
            
        if error_count > 0:
            messages.warning(request, f'有 {error_count} 条记录删除失败，请查看日志了解详情。')
        
        return redirect('inventory:incoming_stock_list')
    
    return redirect('inventory:incoming_stock_list')

@login_required
def incoming_stock_approve(request, pk):
    """确认入库途中产品"""
    incoming_stock = get_object_or_404(IncomingStock, pk=pk)
    
    if incoming_stock.status == '已入库':
        messages.warning(request, '该入库途中产品记录已经确认入库。')
        return redirect('inventory:incoming_stock_list')
    
    if request.method == 'POST':
        try:
            # 创建库存变动记录
            movement = StockMovement.objects.create(
                product=incoming_stock.product,
                warehouse=incoming_stock.warehouse,
                movement_type='IN',
                quantity=incoming_stock.quantity,
                date=timezone.now().date(),
                notes=f'由入库途中记录转入：{incoming_stock.notes}'
            )
            
            # 更新入库途中记录状态
            incoming_stock.status = '已入库'
            incoming_stock.save()
            
            # 不再扣减生产订单剩余数量，因为这个扣减已经在创建入库途中产品时完成
            
            messages.success(request, '入库途中产品已确认入库。')
            
            return redirect('inventory:incoming_stock_list')
        except Exception as e:
            messages.error(request, f'确认入库途中产品时出错: {str(e)}')
            return redirect('inventory:incoming_stock_list')
    
    return render(request, 'inventory/incoming_stock_confirm_approve.html', {'incoming_stock': incoming_stock})

@login_required
def incoming_stock_batch_approve(request):
    """批量确认入库途中产品"""
    if request.method == 'POST':
        selected_ids = request.POST.getlist('selected_items')
        print(f"接收到批量确认入库请求，选中的ID: {selected_ids}")
        
        if not selected_ids:
            messages.error(request, '未选择任何入库途中产品记录。')
            return redirect('inventory:incoming_stock_list')
        
        approved_count = 0
        skipped_count = 0
        error_count = 0
        
        for item_id in selected_ids:
            try:
                incoming_stock = IncomingStock.objects.get(pk=item_id)
                
                # 只处理待入库状态的记录
                if incoming_stock.status != '待入库':
                    print(f"跳过非待入库状态的记录 ID={item_id}, 产品={incoming_stock.product.sku}, 仓库={incoming_stock.warehouse.code}, 状态={incoming_stock.status}")
                    skipped_count += 1
                    continue
                
                # 创建库存变动记录
                movement = StockMovement.objects.create(
                    product=incoming_stock.product,
                    warehouse=incoming_stock.warehouse,
                    movement_type='IN',
                    quantity=incoming_stock.quantity,
                    date=timezone.now().date(),
                    notes=f'由入库途中记录批量转入：{incoming_stock.notes}'
                )
                
                print(f"创建库存变动记录 ID={movement.id}, 产品={incoming_stock.product.sku}, 仓库={incoming_stock.warehouse.code}, 数量={incoming_stock.quantity}")
                
                # 更新入库途中记录状态
                incoming_stock.status = '已入库'
                incoming_stock.save()
                
                # 不再扣减生产订单剩余数量，因为这个扣减已经在创建入库途中产品时完成
                
                approved_count += 1
                print(f"成功确认入库记录 ID={item_id}, 产品={incoming_stock.product.sku}, 仓库={incoming_stock.warehouse.code}")
            except IncomingStock.DoesNotExist:
                error_count += 1
                print(f"入库途中产品记录不存在 ID={item_id}")
            except Exception as e:
                error_count += 1
                print(f"确认入库途中产品记录时出错 ID={item_id}, 错误: {str(e)}")
        
        if approved_count > 0:
            messages.success(request, f'成功确认入库 {approved_count} 条入库途中产品记录。')
        else:
            messages.warning(request, '没有确认入库任何入库途中产品记录。')
        
        if skipped_count > 0:
            messages.info(request, f'跳过了 {skipped_count} 条非待入库状态的记录。')
            
        if error_count > 0:
            messages.warning(request, f'有 {error_count} 条记录处理失败，请查看日志了解详情。')
        
        return redirect('inventory:incoming_stock_list')
    
    return redirect('inventory:incoming_stock_list')

@login_required
def download_incoming_stock_template(request):
    """下载入库途中产品导入模板"""
    # 创建一个新的工作簿
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "入库途中产品导入模板"
    
    # 设置表头
    headers = ['店铺SKU', '仓库代码', '数量', '预计入仓时间(YYYY-MM-DD)', '备注(可选)']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col)
        cell.value = header
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
    
    # 添加示例数据
    example_data = [
        ['SHOP001', 'GA', '100', '2025-03-15', '从供应商A发货'],
        ['SHOP002', 'CA', '50', '2025-03-20', '从供应商B发货']
    ]
    for row, data in enumerate(example_data, 2):
        for col, value in enumerate(data, 1):
            ws.cell(row=row, column=col, value=value)
    
    # 设置列宽
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 20
    
    # 保存到内存中
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    # 返回文件
    response = HttpResponse(buffer.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=incoming_stock_template.xlsx'
    return response

@login_required
def import_incoming_stock(request):
    """导入入库途中产品数据"""
    if request.method == 'POST' and request.FILES.get('file'):
        try:
            excel_file = request.FILES['file']
            df = pd.read_excel(excel_file)
            
            # 验证必要的列是否存在
            required_columns = ['店铺SKU', '仓库代码', '数量', '预计入仓时间(YYYY-MM-DD)']
            for col in required_columns:
                if col not in df.columns:
                    messages.error(request, f'缺少必要的列: {col}')
                    return redirect('inventory:incoming_stock_list')
            
            success_count = 0
            error_count = 0
            error_messages = []
            
            for index, row in df.iterrows():
                try:
                    # 获取产品
                    try:
                        product = Product.objects.get(sku=row['店铺SKU'])
                    except Product.DoesNotExist:
                        error_messages.append(f'第{index+2}行: 店铺SKU {row["店铺SKU"]} 不存在')
                        error_count += 1
                        continue
                    
                    # 获取仓库
                    try:
                        warehouse = Warehouse.objects.get(code=row['仓库代码'])
                    except Warehouse.DoesNotExist:
                        error_messages.append(f'第{index+2}行: 仓库代码 {row["仓库代码"]} 不存在')
                        error_count += 1
                        continue
                    
                    # 验证数量
                    try:
                        quantity = int(row['数量'])
                        if quantity <= 0:
                            error_messages.append(f'第{index+2}行: 数量必须为正整数')
                            error_count += 1
                            continue
                    except:
                        error_messages.append(f'第{index+2}行: 数量必须为整数')
                        error_count += 1
                        continue
                    
                    # 验证并转换日期
                    try:
                        expected_arrival_date = pd.to_datetime(row['预计入仓时间(YYYY-MM-DD)']).date()
                    except:
                        error_messages.append(f'第{index+2}行: 预计入仓时间格式错误')
                        error_count += 1
                        continue
                    
                    # 获取备注（可选）
                    notes = str(row.get('备注(可选)', '')) if pd.notna(row.get('备注(可选)', '')) else ''
                    
                    # 创建入库途中记录
                    incoming_stock = IncomingStock.objects.create(
                        product=product,
                        warehouse=warehouse,
                        quantity=quantity,
                        expected_arrival_date=expected_arrival_date,
                        notes=notes
                    )
                    
                    # 检查是否有相应的生产订单并扣减剩余数量
                    production_orders = ProductionOrder.objects.filter(
                        product=product,
                        remaining_quantity__gt=0
                    )
                    
                    if production_orders.exists():
                        try:
                            # 使用通用的adjust_production_orders函数扣减生产订单数量
                            orders_updated = adjust_production_orders(product, quantity)
                            print(f"导入记录 {index+2}: 扣减生产订单数量完成，更新了{orders_updated}个订单")
                        except Exception as e:
                            print(f"导入记录 {index+2}: 扣减生产订单数量时出错: {str(e)}")
                            error_messages.append(f'第{index+2}行: 扣减生产订单数量时出错 - {str(e)}')
                    else:
                        print(f"导入记录 {index+2}: 产品{product.sku}没有可用的生产订单，无法扣减剩余数量")
                    
                    success_count += 1
                    
                except Exception as e:
                    error_count += 1
                    error_messages.append(f'第{index+2}行: 处理时出错 - {str(e)}')
                    print(f"导入入库途中产品记录时出错: {str(e)}")
            
            if success_count > 0:
                messages.success(request, f'成功导入 {success_count} 条入库途中产品记录。')
            
            if error_count > 0:
                messages.warning(request, f'导入过程中有 {error_count} 条记录出错。')
                for error in error_messages[:10]:  # 只显示前10条错误信息
                    messages.warning(request, error)
                if len(error_messages) > 10:
                    messages.warning(request, f'... 还有 {len(error_messages) - 10} 条错误信息未显示')
            
            return redirect('inventory:incoming_stock_list')
            
        except Exception as e:
            messages.error(request, f'导入文件时出错: {str(e)}')
            return redirect('inventory:incoming_stock_list')
    
    return render(request, 'inventory/incoming_stock_import.html')

@login_required
def export_incoming_stock(request):
    """导出入库途中产品数据"""
    # 创建一个新的工作簿
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "入库途中产品"
    
    # 设置表头
    headers = ['店铺SKU', '产品SKU', '仓库', '仓库代码', '数量', '预计入仓时间', '状态', '备注', '创建时间']
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
    
    # 添加数据
    incoming_stocks = IncomingStock.objects.all()
    for row_num, item in enumerate(incoming_stocks, 2):
        ws.cell(row=row_num, column=1, value=item.product.sku)
        ws.cell(row=row_num, column=2, value=item.product.name)
        ws.cell(row=row_num, column=3, value=item.warehouse.name)
        ws.cell(row=row_num, column=4, value=item.warehouse.code)
        ws.cell(row=row_num, column=5, value=item.quantity)
        ws.cell(row=row_num, column=6, value=item.expected_arrival_date.strftime('%Y-%m-%d'))
        ws.cell(row=row_num, column=7, value=item.status)
        ws.cell(row=row_num, column=8, value=item.notes)
        ws.cell(row=row_num, column=9, value=item.created_at.strftime('%Y-%m-%d %H:%M:%S'))
    
    # 设置列宽
    for col_num in range(1, len(headers) + 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_num)].width = 20
    
    # 保存到内存中
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    # 返回文件
    response = HttpResponse(
        buffer.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=incoming_stock_{}.xlsx'.format(datetime.now().strftime('%Y%m%d%H%M%S'))
    return response

@login_required
def export_stock_movements(request):
    """导出库存变动数据"""
    # 创建一个新的工作簿
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "库存变动"
    
    # 设置表头
    headers = ['日期', '产品SKU', '产品名称', '仓库', '类型', '数量', '备注', '创建时间']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col)
        cell.value = header
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
    
    # 获取所有库存变动记录
    stock_movements = StockMovement.objects.all().order_by('-date', '-created_at')
    
    # 添加数据
    for row, movement in enumerate(stock_movements, 2):
        ws.cell(row=row, column=1, value=movement.date.strftime('%Y-%m-%d'))
        ws.cell(row=row, column=2, value=movement.product.sku)
        ws.cell(row=row, column=3, value=movement.product.name)
        ws.cell(row=row, column=4, value=f"{movement.warehouse.name} ({movement.warehouse.code})")
        ws.cell(row=row, column=5, value='入库' if movement.movement_type == 'IN' else '出库')
        ws.cell(row=row, column=6, value=movement.quantity)
        ws.cell(row=row, column=7, value=movement.notes)
        ws.cell(row=row, column=8, value=movement.created_at.strftime('%Y-%m-%d %H:%M:%S'))
    
    # 设置列宽
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 20
    
    # 保存到内存中
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    # 返回文件
    response = HttpResponse(buffer.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename=stock_movements_{datetime.now().strftime("%Y%m%d%H%M%S")}.xlsx'
    return response

@login_required
def export_historical_stock(request):
    """导出历史库存数据"""
    # 获取日期参数，默认为今天
    date_str = request.GET.get('date', '')
    try:
        if date_str:
            selected_date = datetime.strptime(date_str, '%Y-%m-%d').date()
        else:
            selected_date = timezone.now().date()
    except ValueError:
        selected_date = timezone.now().date()
    
    # 创建一个新的工作簿
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"历史库存_{selected_date.strftime('%Y-%m-%d')}"
    
    # 获取所有仓库
    warehouses = Warehouse.objects.all()
    
    # 设置表头
    headers = ['产品SKU', '产品名称', '重量(lb)', '尺寸(inch)']
    for warehouse in warehouses:
        headers.append(f"{warehouse.name} ({warehouse.code})")
    headers.append('总库存')
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col)
        cell.value = header
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
    
    # 获取所有产品
    products = Product.objects.all()
    
    # 添加数据
    for row, product in enumerate(products, 2):
        ws.cell(row=row, column=1, value=product.sku)
        ws.cell(row=row, column=2, value=product.name)
        ws.cell(row=row, column=3, value=float(product.weight))
        ws.cell(row=row, column=4, value=f"{product.length} × {product.width} × {product.height}")
        
        # 计算每个仓库的历史库存
        col_index = 5
        total_stock = 0
        for warehouse in warehouses:
            # 计算截至所选日期的库存
            stock_in = StockMovement.objects.filter(
                product=product,
                warehouse=warehouse,
                movement_type='IN',
                date__lte=selected_date
            ).aggregate(total=Sum('quantity'))['total'] or 0
            
            stock_out = StockMovement.objects.filter(
                product=product,
                warehouse=warehouse,
                movement_type='OUT',
                date__lte=selected_date
            ).aggregate(total=Sum('quantity'))['total'] or 0
            
            warehouse_stock = stock_in - stock_out
            total_stock += warehouse_stock
            
            ws.cell(row=row, column=col_index, value=warehouse_stock)
            col_index += 1
        
        # 添加总库存
        ws.cell(row=row, column=col_index, value=total_stock)
    
    # 设置列宽
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 20
    
    # 保存到内存中
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    # 返回文件
    response = HttpResponse(buffer.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename=historical_stock_{selected_date.strftime("%Y%m%d")}.xlsx'
    return response

@login_required
def search_products(request):
    """API视图：根据关键字搜索产品"""
    search_term = request.GET.get('term', '')
    products = Product.objects.filter(sku__icontains=search_term)[:20]
    
    results = []
    for product in products:
        results.append({
            'id': product.id,
            'text': f"{product.sku} ({product.name})"
        })
    
    return JsonResponse({'results': results})

@login_required
def incoming_stock_batch_action(request):
    """处理入库途中产品的批量操作（删除和确认入库）"""
    if request.method == 'POST':
        selected_ids = request.POST.getlist('selected_items')
        action = request.POST.get('action', '')
        
        print(f"接收到批量操作请求，操作类型: {action}, 选中的ID: {selected_ids}")
        
        if not selected_ids:
            messages.error(request, '未选择任何入库途中产品记录。')
            return redirect('inventory:incoming_stock_list')
        
        if action == 'delete':
            # 批量删除
            deleted_count = 0
            error_count = 0
            orders_updated_count = 0
            
            for item_id in selected_ids:
                try:
                    item = IncomingStock.objects.get(pk=item_id)
                    item_status = item.status
                    item_sku = item.product.sku if item.product else "未知产品"
                    item_warehouse = item.warehouse.code if item.warehouse else "未知仓库"
                    
                    # 如果是待入库状态，需要恢复生产订单数量
                    if item.status == '待入库':
                        try:
                            # 恢复生产订单数量（传入负数表示增加）
                            orders_updated = adjust_production_orders(item.product, -item.quantity)
                            orders_updated_count += orders_updated
                        except Exception as e:
                            print(f"恢复生产订单数量时出错 ID={item_id}, 错误: {str(e)}")
                    
                    item.delete()
                    deleted_count += 1
                    print(f"成功删除入库途中产品记录 ID={item_id}, 产品={item_sku}, 仓库={item_warehouse}, 状态={item_status}")
                except IncomingStock.DoesNotExist:
                    error_count += 1
                    print(f"入库途中产品记录不存在 ID={item_id}")
                except Exception as e:
                    error_count += 1
                    print(f"删除入库途中产品记录时出错 ID={item_id}, 错误: {str(e)}")
            
            if deleted_count > 0:
                if orders_updated_count > 0:
                    messages.success(request, f'成功删除 {deleted_count} 条入库途中产品记录，并恢复了 {orders_updated_count} 个生产订单的数量。')
                else:
                    messages.success(request, f'成功删除 {deleted_count} 条入库途中产品记录。')
            else:
                messages.warning(request, '没有删除任何入库途中产品记录。')
                
            if error_count > 0:
                messages.warning(request, f'有 {error_count} 条记录删除失败，请查看日志了解详情。')
        
        elif action == 'approve':
            # 批量确认入库
            approved_count = 0
            skipped_count = 0
            error_count = 0
            
            for item_id in selected_ids:
                try:
                    incoming_stock = IncomingStock.objects.get(pk=item_id)
                    
                    # 只处理待入库状态的记录
                    if incoming_stock.status != '待入库':
                        print(f"跳过非待入库状态的记录 ID={item_id}, 产品={incoming_stock.product.sku}, 仓库={incoming_stock.warehouse.code}, 状态={incoming_stock.status}")
                        skipped_count += 1
                        continue
                    
                    # 创建库存变动记录
                    movement = StockMovement.objects.create(
                        product=incoming_stock.product,
                        warehouse=incoming_stock.warehouse,
                        movement_type='IN',
                        quantity=incoming_stock.quantity,
                        date=timezone.now().date(),
                        notes=f'由入库途中记录批量转入：{incoming_stock.notes}'
                    )
                    
                    print(f"创建库存变动记录 ID={movement.id}, 产品={incoming_stock.product.sku}, 仓库={incoming_stock.warehouse.code}, 数量={incoming_stock.quantity}")
                    
                    # 更新入库途中记录状态
                    incoming_stock.status = '已入库'
                    incoming_stock.save()
                    
                    # 不再扣减生产订单剩余数量，因为这个扣减已经在创建入库途中产品时完成
                    
                    approved_count += 1
                    print(f"成功确认入库记录 ID={item_id}, 产品={incoming_stock.product.sku}, 仓库={incoming_stock.warehouse.code}")
                except IncomingStock.DoesNotExist:
                    error_count += 1
                    print(f"入库途中产品记录不存在 ID={item_id}")
                except Exception as e:
                    error_count += 1
                    print(f"确认入库途中产品记录时出错 ID={item_id}, 错误: {str(e)}")
            
            if approved_count > 0:
                messages.success(request, f'成功确认入库 {approved_count} 条入库途中产品记录。')
            else:
                messages.warning(request, '没有确认入库任何入库途中产品记录。')
            
            if skipped_count > 0:
                messages.info(request, f'跳过了 {skipped_count} 条非待入库状态的记录。')
                
            if error_count > 0:
                messages.warning(request, f'有 {error_count} 条记录处理失败，请查看日志了解详情。')
        
        else:
            messages.error(request, '未知的操作类型。')
        
        return redirect('inventory:incoming_stock_list')
    
    return redirect('inventory:incoming_stock_list')

@login_required
def production_order_list(request):
    """生产订单列表视图"""
    production_orders = ProductionOrder.objects.all()
    
    # 计算每个产品的在途数量
    products = Product.objects.all()
    product_incoming_stock = {}
    
    # 获取所有待入库的入库途中产品
    incoming_stocks = IncomingStock.objects.filter(status='待入库')
    
    # 初始化产品在途数据字典
    for product in products:
        product_incoming_stock[product.id] = {
            'GA': 0,
            'CA': 0,
            'total': 0
        }
    
    # 计算每个产品的在途数量
    for incoming in incoming_stocks:
        if incoming.product.id in product_incoming_stock:
            if incoming.warehouse.code == 'GA':
                product_incoming_stock[incoming.product.id]['GA'] += incoming.quantity
            elif incoming.warehouse.code == 'CA':
                product_incoming_stock[incoming.product.id]['CA'] += incoming.quantity
            product_incoming_stock[incoming.product.id]['total'] += incoming.quantity
    
    return render(request, 'inventory/production_order_list.html', {
        'production_orders': production_orders,
        'product_incoming_stock': product_incoming_stock,
    })

@login_required
def production_order_create(request):
    """创建生产订单"""
    if request.method == 'POST':
        form = ProductionOrderForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, '生产订单创建成功！')
            return redirect('inventory:production_order_list')
    else:
        form = ProductionOrderForm()
    
    return render(request, 'inventory/production_order_form.html', {'form': form})

@login_required
def production_order_edit(request, pk):
    """编辑生产订单"""
    production_order = get_object_or_404(ProductionOrder, pk=pk)
    
    if request.method == 'POST':
        form = ProductionOrderForm(request.POST, instance=production_order)
        if form.is_valid():
            # 保存前记录原始数量
            original_quantity = production_order.quantity
            new_quantity = form.cleaned_data['quantity']
            
            # 如果数量发生变化，相应调整剩余数量
            if new_quantity != original_quantity:
                # 计算差值比例
                ratio = new_quantity / original_quantity
                # 按比例调整剩余数量
                production_order.remaining_quantity = int(production_order.remaining_quantity * ratio)
            
            form.save()
            messages.success(request, '生产订单更新成功！')
            return redirect('inventory:production_order_list')
        else:
            form = ProductionOrderForm(instance=production_order)
    
    return render(request, 'inventory/production_order_form.html', {'form': form})

@login_required
def production_order_delete(request, pk):
    """删除生产订单"""
    production_order = get_object_or_404(ProductionOrder, pk=pk)
    
    if request.method == 'POST':
        production_order.delete()
        messages.success(request, '生产订单已成功删除！')
        return redirect('inventory:production_order_list')
    
    return render(request, 'inventory/production_order_confirm_delete.html', {'production_order': production_order})

@login_required
def production_order_batch_delete(request):
    """批量删除生产订单"""
    if request.method == 'POST':
        selected_ids = request.POST.getlist('selected_orders')
        if not selected_ids:
            messages.error(request, '未选择任何生产订单。')
            return redirect('inventory:production_order_list')
        
        deleted_count = 0
        error_count = 0
        
        for order_id in selected_ids:
            try:
                order = ProductionOrder.objects.get(pk=order_id)
                order.delete()
                deleted_count += 1
            except ProductionOrder.DoesNotExist:
                error_count += 1
            except Exception as e:
                error_count += 1
                print(f"删除生产订单时出错 ID={order_id}, 错误: {str(e)}")
        
        if deleted_count > 0:
            messages.success(request, f'成功删除 {deleted_count} 条生产订单。')
        
        if error_count > 0:
            messages.warning(request, f'有 {error_count} 条记录删除失败。')
        
        return redirect('inventory:production_order_list')
    
    return redirect('inventory:production_order_list')

@login_required
def download_production_order_template(request):
    """下载生产订单导入模板"""
    # 创建一个新的工作簿
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "生产订单导入模板"
    
    # 设置表头
    headers = ['店铺SKU', '订单号', '数量']
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    
    # 添加示例数据
    example_data = ['ABC123', 'ORDER-001', '100']
    for col_num, value in enumerate(example_data, 1):
        ws.cell(row=2, column=col_num, value=value)
    
    # 设置列宽
    for col_num in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col_num)].width = 15
    
    # 保存到内存中
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    # 创建响应
    response = HttpResponse(
        buffer.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=production_order_import_template.xlsx'
    
    return response

@login_required
def import_production_orders(request):
    """导入生产订单"""
    if request.method == 'POST' and request.FILES.get('file'):
        try:
            excel_file = request.FILES['file']
            df = pd.read_excel(excel_file)
            
            # 打印列名，用于调试
            print("Excel文件列名:", df.columns.tolist())
            
            # 验证必要的列是否存在
            required_columns = ['店铺SKU', '订单号', '数量']
            for col in required_columns:
                if col not in df.columns:
                    messages.error(request, f'缺少必要的列: {col}')
                    return redirect('inventory:production_order_list')
            
            success_count = 0
            error_count = 0
            error_messages = []
            
            for index, row in df.iterrows():
                try:
                    # 打印当前行数据，用于调试
                    print(f"处理第{index+2}行: {dict(row)}")
                    
                    # 验证店铺SKU
                    sku = str(row['店铺SKU']).strip()
                    if not sku:
                        error_messages.append(f'第{index+2}行: 店铺SKU不能为空')
                        error_count += 1
                        continue
                    
                    # 验证订单号
                    order_number = str(row['订单号']).strip()
                    if not order_number:
                        error_messages.append(f'第{index+2}行: 订单号不能为空')
                        error_count += 1
                        continue
                    
                    # 验证数量
                    try:
                        quantity = int(row['数量'])
                        if quantity <= 0:
                            error_messages.append(f'第{index+2}行: 数量必须为正整数')
                            error_count += 1
                            continue
                    except:
                        error_messages.append(f'第{index+2}行: 数量必须为整数')
                        error_count += 1
                        continue
                    
                    # 查找产品
                    try:
                        product = Product.objects.get(sku=sku)
                    except Product.DoesNotExist:
                        error_messages.append(f'第{index+2}行: 找不到店铺SKU为 {sku} 的产品')
                        error_count += 1
                        continue
                    
                    # 检查是否已存在相同订单号的生产订单
                    existing_order = ProductionOrder.objects.filter(order_number=order_number, product=product).first()
                    if existing_order:
                        error_messages.append(f'第{index+2}行: 产品 {sku} 已存在订单号为 {order_number} 的生产订单')
                        error_count += 1
                        continue
                    
                    # 创建生产订单
                    ProductionOrder.objects.create(
                        product=product,
                        order_number=order_number,
                        quantity=quantity,
                        remaining_quantity=quantity
                    )
                    
                    success_count += 1
                    
                except Exception as e:
                    error_count += 1
                    error_messages.append(f'第{index+2}行: 处理时出错 - {str(e)}')
                    print(f"导入生产订单时出错: {str(e)}")
            
            if success_count > 0:
                messages.success(request, f'成功导入 {success_count} 条生产订单记录。')
            
            if error_count > 0:
                messages.warning(request, f'导入过程中有 {error_count} 条记录出错。')
                for error in error_messages[:10]:  # 只显示前10条错误信息
                    messages.warning(request, error)
                if len(error_messages) > 10:
                    messages.warning(request, f'... 还有 {len(error_messages) - 10} 条错误信息未显示')
            
            return redirect('inventory:production_order_list')
            
        except Exception as e:
            messages.error(request, f'导入文件时出错: {str(e)}')
            return redirect('inventory:production_order_list')
    
    return render(request, 'inventory/production_order_import.html')

@login_required
def export_production_orders(request):
    """导出生产订单"""
    # 创建一个新的工作簿
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "生产订单"
    
    # 设置表头
    headers = ['店铺SKU', '产品SKU', '订单号', '总数量', '剩余数量', '创建时间']
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    
    # 获取所有生产订单
    production_orders = ProductionOrder.objects.all().order_by('-created_at')
    
    # 添加数据
    for row_num, order in enumerate(production_orders, 2):
        ws.cell(row=row_num, column=1, value=order.product.sku)
        ws.cell(row=row_num, column=2, value=order.product.name)
        ws.cell(row=row_num, column=3, value=order.order_number)
        ws.cell(row=row_num, column=4, value=order.quantity)
        ws.cell(row=row_num, column=5, value=order.remaining_quantity)
        ws.cell(row=row_num, column=6, value=order.created_at.strftime('%Y-%m-%d %H:%M'))
    
    # 设置列宽
    for col_num in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col_num)].width = 15
    
    # 保存到内存中
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    # 创建响应
    response = HttpResponse(
        buffer.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=production_orders.xlsx'
    
    return response
